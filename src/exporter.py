# src/exporter.py
from __future__ import annotations

import io
from typing import Optional

import pandas as pd

def export_schedule_xlsx(df_matrix: "pd.DataFrame", year: int, month: int, sheet_name: str = "Cizelge") -> bytes:
    """
    df_matrix DataFrame'ini Excel'e yazar ve bytes döner.
    app.py bunu st.download_button ile indirir.
    """
    xlsx_buf = io.BytesIO()

    # openpyxl ile yaz
    with pd.ExcelWriter(xlsx_buf, engine="openpyxl") as writer:
        (df_matrix if df_matrix is not None else pd.DataFrame()).to_excel(
            writer, index=False, sheet_name=sheet_name
        )

        ws = writer.book[sheet_name]

        # Basit görünüm iyileştirme
        try:
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            # Başlık satırı kalın + ortalı
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Hücreleri wrap + ortala
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Kolon genişliği ayarla
            for col_idx, col in enumerate(ws.columns, start=1):
                max_len = 10
                for cell in col:
                    try:
                        v = "" if cell.value is None else str(cell.value)
                        if len(v) > max_len:
                            max_len = len(v)
                    except Exception:
                        pass
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

            # Üst bilgi (opsiyonel)
            ws.freeze_panes = "A2"
        except Exception:
            pass

    return xlsx_buf.getvalue()
