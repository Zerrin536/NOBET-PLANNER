import streamlit as st
import io
import pandas as pd
from datetime import date, timedelta

from src.db import init_db
from src.staff_repo import (
    add_staff, add_staff_bulk, list_staff, set_staff_active, delete_staff
)
from src.unavailability_repo import (
    add_unavailability_range, list_unavailability, delete_unavailability
)
from src.requests_repo import (
    add_request, list_requests, set_request_status, delete_request, list_approved_requests
)
from src.holidays_repo import (
    add_holidays, list_holidays, delete_holiday
)
from src.calendar_utils import (
    iter_month_days, count_weekdays_excluding_holidays
)
from src.blockers import build_blocked_days_with_type
from src.scheduler import build_required_shifts, SHIFT_HOURS, generate_schedule_hard_min_hours, validate_assignments
from src.assignments_repo import clear_month, insert_assignments, list_month
from src.rules_repo import ensure_rules_table, add_rule, list_rules, set_rule_active, delete_rule
from src.rules_presets import PRESETS, apply_preset

st.set_page_config(page_title="Nöbet Planlayıcı", layout="wide")
init_db()
ensure_rules_table()

st.title("Akıllı Nöbet / Vardiya Planlayıcı")

tab_staff, tab_unav, tab_req, tab_cal, tab_rules, tab_plan = st.tabs(
    ["👩‍⚕️ Personel", "🩺 Rapor / İzin", "📝 İstek Defteri", "📅 Takvim & Tatiller", "⚙️ Kurallar", "📋 Plan"]
)

# -------------------- PERSONEL --------------------
with tab_staff:
    st.subheader("Personel Yönetimi")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### Tek Tek Ekle")
        name = st.text_input("Hemşire adı soyadı", placeholder="Örn: Ayşe Yılmaz", key="staff_name")
        if st.button("Ekle", type="primary", key="staff_add_one"):
            add_staff(name)
            st.success("Eklendi ✅")
            st.rerun()

        st.markdown("---")
        st.markdown("### Toplu Ekle (40 kişi için)")
        bulk = st.text_area(
            "Her satıra 1 isim yaz",
            height=200,
            placeholder="Ayşe Yılmaz\nElif Demir\n...",
            key="staff_bulk"
        )
        if st.button("Toplu Ekle", key="staff_add_bulk"):
            names = [n.strip() for n in bulk.splitlines() if n.strip()]
            count = add_staff_bulk(names)
            st.success(f"{count} kişi eklendi ✅")
            st.rerun()

    with col2:
        st.markdown("### Liste")
        filter_opt = st.radio("Filtre", ["Hepsi", "Aktif", "Pasif"], horizontal=True, key="staff_filter")
        only_active = None
        if filter_opt == "Aktif":
            only_active = True
        elif filter_opt == "Pasif":
            only_active = False

        rows = list_staff(only_active=only_active)

        if not rows:
            st.info("Henüz personel yok.")
        else:
            for r in rows:
                staff_id = int(r["id"])
                full_name = r["full_name"]
                is_active = bool(r["is_active"])

                c1, c2, c3 = st.columns([6, 2, 2])
                with c1:
                    st.write(f"**{full_name}**  (ID: {staff_id})")
                with c2:
                    if is_active:
                        if st.button("Pasif Yap", key=f"deact_{staff_id}"):
                            set_staff_active(staff_id, False)
                            st.rerun()
                    else:
                        if st.button("Aktif Yap", key=f"act_{staff_id}"):
                            set_staff_active(staff_id, True)
                            st.rerun()
                with c3:
                    if st.button("Sil", key=f"del_{staff_id}"):
                        delete_staff(staff_id)
                        st.rerun()

# -------------------- RAPOR / İZİN --------------------
with tab_unav:
    st.subheader("Rapor / Yıllık İzin (Çalışılamayan Günler)")

    staff_rows = list_staff(only_active=True)
    if not staff_rows:
        st.warning("Önce aktif personel eklemelisin.")
    else:
        staff_map = {f'{r["full_name"]} (ID:{r["id"]})': int(r["id"]) for r in staff_rows}
        selected_label = st.selectbox("Personel seç", list(staff_map.keys()), key="u_staff")
        staff_id = staff_map[selected_label]

        c1, c2 = st.columns(2)
        with c1:
            utype = st.selectbox("Tür", ["rapor", "yillik_izin"], key="u_type")
            note = st.text_input("Not (opsiyonel)", placeholder="Örn: okul / sağlık / özel durum", key="u_note")
        with c2:
            start = st.date_input("Başlangıç", value=date.today(), key="u_start")
            end = st.date_input("Bitiş", value=date.today(), key="u_end")

        if end < start:
            st.error("Bitiş tarihi başlangıçtan önce olamaz.")
        else:
            days = []
            cur = start
            while cur <= end:
                days.append(cur.isoformat())
                cur += timedelta(days=1)

            if st.button("Kaydet", type="primary", key="u_save"):
                add_unavailability_range(staff_id, days, utype, note)
                st.success(f"{len(days)} gün kaydedildi ✅")
                st.rerun()

        st.markdown("---")
        st.markdown("### Kayıtlar")
        filt = st.checkbox("Sadece seçili personeli göster", value=True, key="u_filt")
        rows = list_unavailability(staff_id if filt else None)

        if not rows:
            st.info("Kayıt yok.")
        else:
            for r in rows:
                rid = int(r["id"])
                extra = ((" | " + r["note"]) if r["note"] else "")
                st.write(f'**{r["date"]}** — {r["full_name"]} — `{r["type"]}`{extra}')
                if st.button("Sil", key=f"unav_del_{rid}"):
                    delete_unavailability(rid)
                    st.rerun()

# -------------------- İSTEK DEFTERİ --------------------
with tab_req:
    st.subheader("İstek Defteri (Okul / Eğitim / Çocuk / Özel Durum)")

    staff_rows = list_staff(only_active=True)
    if not staff_rows:
        st.warning("Önce aktif personel eklemelisin.")
    else:
        staff_map = {f'{r["full_name"]} (ID:{r["id"]})': int(r["id"]) for r in staff_rows}
        selected_label = st.selectbox("Personel seç (şimdilik login yerine)", list(staff_map.keys()), key="req_staff")
        staff_id = staff_map[selected_label]

        req_kind = st.selectbox("İstek tipi", ["HARD", "SOFT"], index=0, key="req_kind")
        req_day = st.date_input("İstek günü", value=date.today() + timedelta(days=14), key="req_day")
        req_note = st.text_area(
            "İstek notu",
            height=120,
            placeholder="Örn: Okul var, o gün boş olsun / Eğitim var / Çocuk hastane randevusu…",
            key="req_note"
        )

        days_ahead = (req_day - date.today()).days
        if days_ahead < 14:
            st.warning(
                f"Bu istek {days_ahead} gün sonra. Kural: en az 14 gün önceden bildirim. "
                "Yine de kaydedilebilir, yönetici karar verir."
            )

        if st.button("İstek Kaydet", type="primary", key="req_save"):
            add_request(staff_id, req_day.isoformat(), req_note, req_kind)
            st.success("İstek kaydedildi ✅")
            st.rerun()

        st.markdown("---")
        st.markdown("### Yönetici Görünümü (İstek Listesi)")

        status_label_map = {"Beklemede": "pending", "Onaylandı": "approved", "Reddedildi": "rejected"}
        status_tr_map = {"pending": "Beklemede", "approved": "Onaylandı", "rejected": "Reddedildi"}

        filter_status = st.radio(
            "Durum filtresi",
            ["Hepsi", "Beklemede", "Onaylandı", "Reddedildi"],
            horizontal=True,
            key="req_filter_status"
        )
        status = None if filter_status == "Hepsi" else status_label_map[filter_status]
        reqs = list_requests(status=status)

        if not reqs:
            st.info("İstek yok.")
        else:
            for r in reqs:
                rid = int(r["id"])
                kind = (r.get("request_kind") or "HARD").upper()
                st.write(
                    f'**{r["date"]}** — {r["full_name"]} — '
                    f'**{status_tr_map.get(r["status"], r["status"])}**  |  Tip: `{kind}`'
                )
                st.caption(f'{r["note"]}  |  Oluşturma: {r["created_at"]}')

                c1, c2, c3 = st.columns([1, 1, 1])
                with c1:
                    if st.button("Onayla", key=f"req_ok_{rid}"):
                        set_request_status(rid, "approved")
                        st.rerun()
                with c2:
                    if st.button("Reddet", key=f"req_no_{rid}"):
                        set_request_status(rid, "rejected")
                        st.rerun()
                with c3:
                    if st.button("Sil", key=f"req_del_{rid}"):
                        delete_request(rid)
                        st.rerun()

        st.info("Not: Onaylı HARD istekler planlamada 'kesin boş' (hard). Onaylı SOFT istekler ise mümkünse boş bırakılır.")

# -------------------- TAKVİM & TATİLLER --------------------
with tab_cal:
    st.subheader("Takvim & Resmî Tatiller")

    today = date.today()
    c1, c2 = st.columns(2)
    with c1:
        year = st.number_input("Yıl", min_value=2020, max_value=2100, value=today.year, step=1, key="cal_year")
    with c2:
        month = st.selectbox("Ay", list(range(1, 13)), index=today.month - 1, key="cal_month")

    days = iter_month_days(int(year), int(month))
    st.caption(f"Seçilen ay gün sayısı: **{len(days)}**")

    holiday_list = list_holidays()
    holiday_set = set(holiday_list)

    st.markdown("### Ay günleri (Tatil seç)")
    options = [d.iso for d in days]
    default_selected = [d for d in options if d in holiday_set]
    selected_holidays = st.multiselect("Resmî tatiller", options=options, default=default_selected, key="cal_hols")

    if st.button("Tatil Kaydet", type="primary", key="cal_save"):
        for d in default_selected:
            delete_holiday(d)
        add_holidays(selected_holidays)
        st.success("Tatiller güncellendi ✅")
        st.rerun()

    st.markdown("---")
    weekday_count = count_weekdays_excluding_holidays(int(year), int(month), set(selected_holidays))
    min_month_hours = weekday_count * 8

    st.markdown("### Minimum aylık mesai hesabı")
    st.write(f"**Hafta içi gün sayısı (tatiller hariç):** {weekday_count}")
    st.write(f"**Minimum aylık mesai (hard):** {weekday_count} × 8 = **{min_month_hours} saat**")

# -------------------- KURALLAR (EKLE/SİL) --------------------
with tab_rules:
    st.subheader("Kural Yönetimi (PrevShift → NextShift yasak)")

    st.info("Buraya sadece 'YASAK' kuralları ekliyoruz. Serbest olanları eklemen gerekmez.")
    # --- KURAL SETLERI (Varsayılan / Katı / Esnek) ---
    st.markdown("### 🧩 Kural Seti Seç (Hazır Preset)")

    preset_name = st.selectbox(
        "Preset",
        list(PRESETS.keys()),
        index=0,
        key="preset_name"
    )

    # Katı seçilince varsayılan daha kuralcı gelsin
    default_deactivate = True if preset_name == "Katı" else False
    deactivate_others = st.checkbox(
        "Preset dışındaki aktif kuralları pasif yap (temiz set)",
        value=default_deactivate,
        key="preset_deactivate_others"
    )

    # Preset önizleme
    try:
        import pandas as pd
        df_p = pd.DataFrame(PRESETS.get(preset_name, []))
        if not df_p.empty:
            st.dataframe(df_p, width="stretch", height=180)
    except Exception:
        pass

    if st.button("✅ Preset'i Uygula", type="primary", key="apply_preset_btn"):
        touched = apply_preset(preset_name, deactivate_others=deactivate_others)
        st.success(f"Preset uygulandı ✅ (etkilenen kural: {touched})")
        st.rerun()

    st.markdown("---")
    # --- /KURAL SETLERI ---


    c1, c2, c3 = st.columns([2, 2, 4])
    with c1:
        prev_type = st.selectbox(
            "Dün (Prev)",
            ["DAY", "NIGHT", "D24", "RAPOR", "YILLIK_IZIN", "ANY"],
            index=1,
            key="rule_prev"
        )
    with c2:
        next_type = st.selectbox("Bugün (Next)", ["DAY", "NIGHT", "D24", "ANY"], index=3, key="rule_next")
    with c3:
        apply_day = st.selectbox("Hangi günlerde geçerli?", ["ANY", "WEEKDAY", "WEEKEND"], index=0, key="rule_apply_day")
        note = st.text_input("Not (opsiyonel)", placeholder="Örn: 24 sonrası ertesi gün çalışma yok", key="rule_note")

    if st.button("Kural Ekle", type="primary", key="rule_add"):
        add_rule(prev_type, next_type, apply_day, note)
        st.success("Kural eklendi ✅")
        st.rerun()

    st.markdown("---")
    st.markdown("### 📌 Kural Listesi (Arama + Filtre + Gruplu)")

    all_rows_for_count = list_rules(active_only=None)
    active_count = sum(1 for r in all_rows_for_count if bool(r.get("is_active", 1)))
    st.caption(f"Toplam kural: **{len(all_rows_for_count)}** | Aktif: **{active_count}** | Pasif: **{len(all_rows_for_count) - active_count}**")

    f1, f2, f3 = st.columns([2, 2, 6])
    with f1:
        filt = st.radio("Filtre", ["Hepsi", "Aktif", "Pasif"], horizontal=True, key="rules_filter_rules_tab_v2")
    with f2:
        group_mode = st.checkbox("Gün tipine göre grupla", value=True, key="rules_group_mode_v2")
    with f3:
        q = st.text_input("Kural ara (prev/next/not/gün)", value="", placeholder="örn: gece, 24, izin, hafta içi…", key="rules_search_v2")

    if filt == "Aktif":
        rows = list_rules(active_only=True)
    elif filt == "Pasif":
        rows = list_rules(active_only=False)
    else:
        rows = list_rules(active_only=None)

    if q.strip():
        qq = q.strip().lower()
        def _match_row(r):
            s = f'{r.get("prev_type","")} {r.get("next_type","")} {r.get("apply_day","")} {r.get("note","")}'.lower()
            return qq in s
        rows = [r for r in rows if _match_row(r)]

    if not rows:
        st.info("Filtre/arama sonucunda kural bulunamadı.")
    else:
        day_tr = {"ANY": "Her gün", "WEEKDAY": "Hafta içi", "WEEKEND": "Hafta sonu"}

        def _render_rows(rrs, prefix=""):
            for r in rrs:
                rid = int(r["id"])
                is_active = bool(r.get("is_active", 1))
                label = "✅ Aktif" if is_active else "⛔ Pasif"
                ad = r.get("apply_day", "ANY")
                ad_tr = day_tr.get(ad, ad)

                st.write(
                    f'**#{rid}**  {label}  —  `{r.get("prev_type")} -> {r.get("next_type")}`  '
                    f'(Gün: **{ad_tr}**)  {(" | " + r["note"]) if r.get("note") else ""}'
                )

                b1, b2, b3 = st.columns([1, 1, 3])
                with b1:
                    if is_active:
                        if st.button("Pasif Yap", key=f"{prefix}rule_off_{rid}_v2"):
                            set_rule_active(rid, False); st.rerun()
                    else:
                        if st.button("Aktif Yap", key=f"{prefix}rule_on_{rid}_v2"):
                            set_rule_active(rid, True); st.rerun()
                with b2:
                    if st.button("Sil", key=f"{prefix}rule_del_{rid}_v2"):
                        delete_rule(rid); st.rerun()
                with b3:
                    st.caption(r.get("created_at", ""))

        if group_mode:
            any_rows = [r for r in rows if (r.get("apply_day","ANY") == "ANY")]
            wd_rows  = [r for r in rows if (r.get("apply_day","ANY") == "WEEKDAY")]
            we_rows  = [r for r in rows if (r.get("apply_day","ANY") == "WEEKEND")]

            with st.expander(f"📅 Her gün ({len(any_rows)})", expanded=True):
                _render_rows(any_rows, prefix="any_")
            with st.expander(f"🏙️ Hafta içi ({len(wd_rows)})", expanded=False):
                _render_rows(wd_rows, prefix="wd_")
            with st.expander(f"🌙 Hafta sonu ({len(we_rows)})", expanded=False):
                _render_rows(we_rows, prefix="we_")
        else:
            _render_rows(rows, prefix="flat_")



    st.markdown("---")

with tab_plan:
    st.subheader("Plan (Basit v0 + HARD min mesai + Kurallar + İstekler)")

    today = date.today()
    c1, c2 = st.columns(2)
    with c1:
        year = st.number_input("Yıl", min_value=2020, max_value=2100, value=today.year, step=1, key="p_year")
    with c2:
        month = st.selectbox("Ay", list(range(1, 13)), index=today.month - 1, key="p_month")

    staff_rows = list_staff(only_active=True)
    if not staff_rows:
        st.warning("Önce aktif personel eklemelisin.")
    else:
        staff_ids = [int(r["id"]) for r in staff_rows]
        staff_name_by_id = {int(r["id"]): r["full_name"] for r in staff_rows}

        required = build_required_shifts(int(year), int(month))
        st.caption(f"Bu ay toplam slot: **{len(required)}** (hafta içi 24 kişi/gün, hafta sonu 12 kişi/gün)")

        holiday_set = set(list_holidays())
        weekday_count = count_weekdays_excluding_holidays(int(year), int(month), holiday_set)
        min_required_hours = weekday_count * 8
        st.info(f"Hard kural: Her çalışan en az **{min_required_hours} saat** çalışmalı.")

        # ---- Onaylı istekler (HARD/SOFT) ----
        st.markdown("---")
        st.markdown("### ✅ Onaylı İstekler (HARD / SOFT)")
        approved = list_approved_requests(int(year), int(month))
        if not approved:
            st.info("Bu ay için onaylı istek yok.")
        else:
            for r in approved:
                kind = (r.get("request_kind") or "HARD").upper()
                st.write(f'**{r["date"]}** — {r["full_name"]} — Tip: `{kind}`')
                if r.get("note"):
                    st.caption(r["note"])

        # ---- SOFT isteğe rağmen atananlar (şeffaflık) ----
        st.markdown("### ⚠️ SOFT isteğe rağmen atananlar")
        plan_rows_now = list_month(int(year), int(month))
        if not plan_rows_now:
            st.info("Önce plan üretince burada SOFT çatışmalarını göstereceğim.")
        else:
            assigned_set = {(r["date"], int(r["staff_id"])) for r in plan_rows_now}
            soft_conflicts = []
            for r in approved:
                kind = (r.get("request_kind") or "HARD").upper()
                if kind == "SOFT":
                    if (r["date"], int(r["staff_id"])) in assigned_set:
                        soft_conflicts.append(r)

            if not soft_conflicts:
                st.success("SOFT isteklerle çakışan atama yok ✅ (mümkün olduğunca boş bırakıldı)")
            else:
                st.warning(f"SOFT isteğe rağmen atanan kişi sayısı: {len(soft_conflicts)}")
                for r in soft_conflicts:
                    st.write(f'**{r["date"]}** — {r["full_name"]} — Tip: `SOFT` (ama atanmış)')
                    if r.get("note"):
                        st.caption(r["note"])

        transition_rules = list_rules(active_only=True)
        st.write(f"Aktif geçiş kuralı sayısı: **{len(transition_rules)}**")

        st.markdown("---")
        st.markdown("### DOGRULAMA (VALID/INVALID)")

        _val = st.session_state.get("last_validation")
        if not _val:
            st.info("Henüz doğrulama yok. Plan üretince burada otomatik kontrol göreceksin.")
        else:
            summary = _val.get("summary", {})
            violations = _val.get("violations", [])
            deficits = _val.get("deficits", [])
            unfilled_n = _val.get("unfilled_count", None)

            hard_ok = bool(summary.get("hard_ok", True))
            min_ok  = bool(summary.get("min_hours_ok", True))
            unfilled_ok = (unfilled_n == 0) if isinstance(unfilled_n, int) else None

            c1, c2, c3 = st.columns(3)
            with c1:
                st.success("Hard ihlal yok ✅" if hard_ok else "Hard ihlal var ❌")
            with c2:
                st.success("Min saat tuttu ✅" if min_ok else "Min saat tutmadi ❌")
            with c3:
                if unfilled_ok is None:
                    st.info("Unfilled bilgisi yok")
                else:
                    st.success("Unfilled yok ✅" if unfilled_ok else "Unfilled var ❌")

            if deficits:
                st.warning("Min saat altinda kalan ID: " + ", ".join(str(x) for x in deficits))

            if violations:
                import pandas as pd
                dfv = pd.DataFrame(violations)
                st.dataframe(dfv, width="stretch", height=260)
            else:
                st.caption("Ihlal listesi: bos")


        if st.button("Plan Üret (Hard min)", type="primary", key="plan_btn"):
            blocked_any, blocked_type, soft_avoid = build_blocked_days_with_type(int(year), int(month))

            # kişi-bazlı min saat: rapor/izin (hafta içi) kadar 8 saat düş
            from datetime import date as _date
            holiday_set_local = set(list_holidays())
            off_weekday = {}
            for sid in staff_ids:
                cnt = 0
                for d_iso, t in (blocked_type.get(sid, {}) or {}).items():
                    try:
                        dt = _date.fromisoformat(str(d_iso)[:10])
                        if dt.weekday() < 5 and str(d_iso)[:10] not in holiday_set_local:
                            # rapor veya yıllık izin olan hafta içi gün -> min saatten düş
                            if t in ('rapor','yillik_izin'):
                                cnt += 1
                    except Exception:
                        pass
                off_weekday[sid] = cnt
            min_by_staff = {sid: max(0, int(min_required_hours) - off_weekday.get(sid,0)*8) for sid in staff_ids}
            assignments, unfilled, unfilled_debug, hours, swaps = generate_schedule_hard_min_hours(
                int(year), int(month), staff_ids, blocked_any, min_by_staff,
                transition_rules=transition_rules,
                blocked_type=blocked_type,
                soft_avoid=soft_avoid
            )

            # --- VALIDATION hesapla (kalıcı) ---
            try:
                v_summary, v_violations, v_deficits = validate_assignments(
                    int(year), int(month),
                    assignments,
                    staff_ids,
                    blocked_any,
                    min_required_hours,
                    transition_rules=transition_rules,
                    blocked_type=blocked_type
                )
                st.session_state["last_validation"] = {
                    "summary": v_summary,
                    "violations": v_violations,
                    "deficits": v_deficits,
                    "unfilled_count": len(unfilled) if unfilled is not None else None,
                }
            except Exception as e:
                st.session_state["last_validation"] = {
                    "summary": {"hard_ok": False, "min_hours_ok": False},
                    "violations": [{"type":"VALIDATION_ERROR","date":"","shift_type":"","staff_id":-1,"detail":str(e)}],
                    "deficits": [],
                    "unfilled_count": len(unfilled) if unfilled is not None else None,
                }
            # --- /VALIDATION ---

            clear_month(int(year), int(month))
            if assignments and isinstance(assignments[0], (tuple, list)):
                assignments = [{"date": a[0], "shift_type": a[1], "staff_id": a[2]} for a in assignments]
            insert_assignments(assignments)

            st.markdown("---")

            # ✅ Unfilled raporu: ekranda kaybolmasın diye session_state'e yaz
            st.session_state["last_unfilled"] = unfilled
            st.session_state["last_unfilled_debug"] = unfilled_debug if 'unfilled_debug' in locals() else []
            st.session_state["last_unfilled_year"] = int(year)
            st.session_state["last_unfilled_month"] = int(month)
            deficits = [sid for sid in staff_ids if hours.get(sid, 0) < int(min_by_staff.get(sid, min_required_hours))]

            if deficits:
                st.error(
                    f"Hard min hedefi tam sağlanamadı. Eksi kalan kişi sayısı: {len(deficits)}. "
                    "Bu, mevcut kurallar ile %100 mümkün olmayabilir."
                )
            else:
                st.success(f"Plan kaydedildi ✅ (Dengeleme swap sayısı: {swaps})")

            st.rerun()

        st.markdown("---")

        # ================== 🚫 Dolmayan Slotlar (Neden Raporu) ==================
        st.markdown("---")
        st.markdown("### 🚫 Dolmayan Slotlar (Neden Raporu)")

        last_y = st.session_state.get("last_unfilled_year")
        last_m = st.session_state.get("last_unfilled_month")
        last_unfilled = st.session_state.get("last_unfilled")
        last_debug = st.session_state.get("last_unfilled_debug")

        if last_unfilled is None or last_y != int(year) or last_m != int(month):
            st.info("Plan ürettikten sonra burada dolmayan slotlar ve nedeni görünecek.")
        else:
            if not last_unfilled:
                st.success("Bu ay dolmayan slot yok ✅")
            else:
                st.warning(f"Doldurulamayan slot: {len(last_unfilled)} (kural çakışması / personel yetersizliği olabilir)")

                # Debug varsa özet tablo
                if last_debug:
                    df_unfilled = pd.DataFrame(last_debug)
                    g = df_unfilled.groupby(["date", "shift_type"], as_index=False).agg({
                        "need": "sum",
                        "assigned": "sum",
                        "missing": "sum",
                        "reason": "first",
                    }).sort_values(["date", "shift_type"])

                    st.dataframe(g, width="stretch", height=320)
                    st.download_button(
                        "📄 Dolmayan slot raporu (CSV)",
                        data=g.to_csv(index=False).encode("utf-8"),
                        file_name=f"unfilled_{int(year)}_{int(month):02d}.csv",
                        mime="text/csv",
                        key="dl_unfilled_csv"
                    )
                else:
                    df_unfilled = pd.DataFrame(last_unfilled)
                    if "day" in df_unfilled.columns:
                        df_unfilled.rename(columns={"day": "date"}, inplace=True)
                    st.dataframe(df_unfilled, width="stretch", height=260)
        # ================== /🚫 Dolmayan Slotlar ==================

                st.markdown("### Plan çıktısı (gün gün)")

        rows = list_month(int(year), int(month))
        if not rows:
            st.info("Bu ay için plan yok. 'Plan Üret' butonuna bas.")
        else:
            # ===== 4+ GÜN BOŞLUK RAPORU =====
            st.markdown("---")
            st.markdown("### 💤 4+ Gün Boşluk Raporu")

            off_runs = []
            for sid in staff_ids:
                run_start = None
                run_len = 0
                has_blocked = False

                # (AUTO-FIX) day_isos tanımı (NameError fix)
                day_infos = iter_month_days(int(year), int(month))
                day_isos = [d.iso for d in day_infos]

                for d_iso in day_isos:
                    # (AUTO-FIX) cell map: 4+ gün boşluk raporu için hızlı lookup
                    try:
                        cell = {(int(r.get('staff_id')), r.get('date')): (r.get('shift_type') or '') for r in rows if r.get('date')}
                    except Exception:
                        cell = {}

                    worked_flag = (sid, d_iso) in cell
                    # (AUTO-FIX) blocked_type: 4+ gün boşluk raporu bloğu için
                    if 'blocked_type' not in locals():
                        blocked_type = {}

                    bt = blocked_type.get(sid, {}).get(d_iso)
                    is_off = (not worked_flag)

                    if is_off:
                        if run_start is None:
                            run_start = d_iso
                            run_len = 1
                            has_blocked = bool(bt)
                        else:
                            run_len += 1
                            if bt:
                                has_blocked = True
                    else:
                        if run_start is not None and run_len >= 4:
                            off_runs.append({
                                "Personel": staff_name_by_id.get(sid, f"ID:{sid}"),
                                "ID": sid,
                                "Baslangic": run_start,
                                "Bitis": (date.fromisoformat(d_iso) - timedelta(days=1)).isoformat(),
                                "Gun": run_len,
                                "AraliktaRaporIzinVar": "Evet" if has_blocked else "Hayır",
                            })
                        run_start = None
                        run_len = 0
                        has_blocked = False

                if run_start is not None and run_len >= 4:
                    off_runs.append({
                        "Personel": staff_name_by_id.get(sid, f"ID:{sid}"),
                        "ID": sid,
                        "Baslangic": run_start,
                        "Bitis": day_isos[-1],
                        "Gun": run_len,
                        "AraliktaRaporIzinVar": "Evet" if has_blocked else "Hayır",
                    })

            if not off_runs:
                st.success("4+ gün boşluk yok ✅")
            else:
                df_off = pd.DataFrame(off_runs).sort_values(["Gun", "Personel"], ascending=[False, True])
                st.warning(f"4+ gün boşluk bulunan kayıt sayısı: {len(df_off)}")
                st.dataframe(df_off, width="stretch", height=260)
                st.download_button(
                    "📄 4+ gün boşluk raporu (CSV)",
                    data=df_off.to_csv(index=False).encode("utf-8"),
                    file_name=f"bosluk_4plus_{int(year)}_{int(month):02d}.csv",
                    mime="text/csv",
                    key="dl_gap_csv"
                )
            # ===== /4+ GÜN BOŞLUK RAPORU =====

            st.markdown("---")
            
            
            st.markdown("### 📊 Aylık Çizelge (Sadece Yazı Rengi)")

            blocked_type = {}
            unav_count = 0
            yyyymm = f"{int(year)}-{int(month):02d}-"

            name_to_id_local = {v: k for k, v in staff_name_by_id.items()}

            def _as_dict(x):
                try:
                    return dict(x)
                except Exception:
                    return x if isinstance(x, dict) else {}

            try:
                unav_rows = list_unavailability(None)

                for ur0 in unav_rows:
                    ur = _as_dict(ur0)
                    d = ur.get("date") or ur.get("day")
                    if not d:
                        continue
                    d = str(d)
                    d = d.split(" ")[0].split("T")[0]

                    if not d.startswith(yyyymm):
                        continue

                    sid = ur.get("staff_id")
                    if sid is None:
                        sid = name_to_id_local.get(ur.get("full_name"))
                    if sid is None:
                        continue
                    sid = int(sid)

                    t = ur.get("type") or ur.get("utype")
                    if t not in ("rapor", "yillik_izin"):
                        continue

                    blocked_type.setdefault(sid, {})[d] = t
                    unav_count += 1

            except Exception as e:
                st.warning(f"Rapor/izin okunamadı: {e}")
                blocked_type = {}
                unav_count = 0

            st.caption(f"Bu ay rapor/izin kaydı: {unav_count}")
            
            name_to_id = {v: k for k, v in staff_name_by_id.items()}

            day_infos = iter_month_days(int(year), int(month))
            day_isos = [d.iso for d in day_infos]
            day_cols = [str(int(d.iso.split("-")[2])).zfill(2) for d in day_infos]

            staff_df = pd.DataFrame(
                [{"Personel": staff_name_by_id.get(sid, f"ID:{sid}"), "ID": sid} for sid in staff_ids]
            ).sort_values(["Personel", "ID"]).reset_index(drop=True)

            cell = {}
            for r in rows:
                sid = r.get("staff_id")
                if sid is None:
                    sid = name_to_id.get(r.get("full_name"))
                if sid is None:
                    continue
                sid = int(sid)

                d_iso = r.get("date")
                stype = r.get("shift_type")
                if not d_iso or not stype:
                    continue

                key = (sid, d_iso)
                if key in cell and stype not in cell[key].split("+"):
                    cell[key] = cell[key] + "+" + stype
                else:
                    cell.setdefault(key, stype)

            matrix_rows = []
            for _, rr in staff_df.iterrows():
                sid = int(rr["ID"])
                row = {"Personel": rr["Personel"], "ID": sid}

                for d_iso, dcol in zip(day_isos, day_cols):
                    bt = blocked_type.get(sid, {}).get(d_iso)
                    if bt == "rapor":
                        row[dcol] = "R"
                        continue
                    if bt == "yillik_izin":
                        row[dcol] = "İ"
                        continue

                    stype = cell.get((sid, d_iso), "")
                    if stype.startswith("DAY"):
                        row[dcol] = "8"
                    elif stype.startswith("NIGHT"):
                        row[dcol] = "16"
                    elif stype.startswith("D24"):
                        row[dcol] = "24"
                    else:
                        row[dcol] = ""
                matrix_rows.append(row)

            df_matrix = pd.DataFrame(matrix_rows)

            aciklama = []
            for sid in df_matrix["ID"].tolist():
                sid = int(sid)
                mp = blocked_type.get(sid, {})
                rapor = sum(1 for _, t in mp.items() if t == "rapor")
                izin  = sum(1 for _, t in mp.items() if t == "yillik_izin")
                parts = []
                if rapor:
                    parts.append(f"{rapor} gün raporlu")
                if izin:
                    parts.append(f"{izin} gün izinli")
                aciklama.append(" | ".join(parts) if parts else "")
            df_matrix["Açıklama"] = aciklama

            min_required_hours = weekday_count * 8

            HOURS_MAP = {"DAY": 8, "NIGHT": 16, "D24": 24}

            worked = {sid: 0 for sid in staff_ids}
            for rr in rows:
                sid = rr.get("staff_id")
                if sid is None:
                    sid = name_to_id.get(rr.get("full_name"))
                if sid is None:
                    continue
                sid = int(sid)

                stype = (rr.get("shift_type") or "")
                parts = str(stype).split("+")
                for part in parts:
                    part = part.strip()
                    if part.startswith("DAY"):
                        worked[sid] += HOURS_MAP["DAY"]
                    elif part.startswith("NIGHT"):
                        worked[sid] += HOURS_MAP["NIGHT"]
                    elif part.startswith("D24"):
                        worked[sid] += HOURS_MAP["D24"]

            # kişi-bazlı MinSaat (çizelge): rapor/izin hafta içi gün * 8 düş
            from datetime import date as _date
            holiday_set_local = set(list_holidays())
            min_by_staff_matrix = {}
            for _sid in staff_ids:
                _cnt = 0
                for _d_iso, _t in (blocked_type.get(int(_sid), {}) or {}).items():
                    try:
                        _dd = _date.fromisoformat(str(_d_iso)[:10])
                        if _dd.weekday() < 5 and str(_d_iso)[:10] not in holiday_set_local:
                            if _t in ('rapor','yillik_izin'):
                                _cnt += 1
                    except Exception:
                        pass
                min_by_staff_matrix[int(_sid)] = max(0, int(min_required_hours) - _cnt*8)
            df_matrix["MinSaat"] = df_matrix["ID"].map(lambda x: int(min_by_staff_matrix.get(int(x), min_required_hours)))
            df_matrix["CalistigiSaat"] = df_matrix["ID"].map(lambda x: worked.get(int(x), 0))
            df_matrix["Fark"] = df_matrix["CalistigiSaat"] - df_matrix["MinSaat"]

            ordered_cols = ["Personel", "ID"] + day_cols + ["MinSaat", "CalistigiSaat", "Fark", "Açıklama"]
            df_matrix = df_matrix[ordered_cols]

            COLOR_DAY   = "#2F75B5"
            COLOR_NIGHT = "#C9A100"
            COLOR_D24   = "#2E8B57"
            COLOR_BLOCK = "#C00000"

            def style_text_only(data: pd.DataFrame):
                styles = pd.DataFrame("", index=data.index, columns=data.columns)
                for c in day_cols:
                    for i in data.index:
                        v = str(data.loc[i, c] or "")
                        if v == "8":
                            styles.loc[i, c] = f"color: {COLOR_DAY};"
                        elif v == "16":
                            styles.loc[i, c] = f"color: {COLOR_NIGHT};"
                        elif v == "24":
                            styles.loc[i, c] = f"color: {COLOR_D24};"
                        elif v in ("R", "İ"):
                            styles.loc[i, c] = f"color: {COLOR_BLOCK};"
                return styles

            styler = df_matrix.style.apply(style_text_only, axis=None)
            st.dataframe(styler, width='stretch', height=680)

            st.markdown("### ⬇️ Çizelgeyi İndir (CSV / Excel)")
            st.download_button(
                "📄 CSV indir (Çizelge)",
                data=df_matrix.to_csv(index=False).encode("utf-8"),
                file_name=f"cizelge_{int(year)}_{int(month):02d}.csv",
                mime="text/csv",
                key="dl_matrix_csv"
            )

            xlsx_buf = io.BytesIO()
            try:
                from openpyxl.styles import Font
                from openpyxl.utils import get_column_letter

                FONT_DAY   = Font(color="2F75B5")
                FONT_NIGHT = Font(color="C9A100")
                FONT_D24   = Font(color="2E8B57")
                FONT_BLOCK = Font(color="C00000")

                with pd.ExcelWriter(xlsx_buf, engine="openpyxl") as writer:
                    df_matrix.to_excel(writer, index=False, sheet_name="Cizelge")
                    ws = writer.sheets["Cizelge"]

                    headers = [cell.value for cell in ws[1]]
                    day_headers = [h for h in headers if isinstance(h, str) and len(h) == 2 and h.isdigit()]
                    day_cols_idx = [headers.index(h) + 1 for h in day_headers]

                    for r in range(2, ws.max_row + 1):
                        for c in day_cols_idx:
                            v = ws.cell(row=r, column=c).value
                            if v is None:
                                continue
                            v = str(v).strip()
                            if v == "8":
                                ws.cell(row=r, column=c).font = FONT_DAY
                            elif v == "16":
                                ws.cell(row=r, column=c).font = FONT_NIGHT
                            elif v == "24":
                                ws.cell(row=r, column=c).font = FONT_D24
                            elif v in ("R", "İ"):
                                ws.cell(row=r, column=c).font = FONT_BLOCK

                    for col_idx in range(1, ws.max_column + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 10

                xlsx_buf.seek(0)
                st.download_button(
                    "📊 Excel indir (.xlsx)",
                    data=xlsx_buf,
                    file_name=f"cizelge_{int(year)}_{int(month):02d}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_matrix_xlsx"
                )
            except Exception as e:
                st.warning(f"Excel export çalışmadı: {str(e)} (CSV her zaman çalışır)")

            st.markdown("---")
            st.caption("Not: Aşağıdaki eski uzun listeyi artık göstermiyoruz; çizelge yukarıda.")
